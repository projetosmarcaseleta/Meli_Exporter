"""
app.py – Interface web Flask para Relatorios Meli
Gera planilha Excel com: SKU | MLB | TIPO ANÚNCIO | CATÁLOGO | QTD VENDIDA
"""

import os
import io
import uuid
import json

from flask import Flask, render_template, request, jsonify, send_file, redirect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from api import validate_token
from exporter import process_mlbs, HEADERS

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "relatorios-meli-dev")
PUBLIC_EXPORT_URL = os.environ.get("PUBLIC_EXPORT_URL", "https://app.marcaseleta.shop/export")


def _write_excel(rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório ML"

    header_fill = PatternFill("solid", fgColor="FFE600")
    header_font = Font(bold=True, color="000000")

    for col_idx, header in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajusta largura das colunas automaticamente
    col_widths = {"SKU": 20, "MLB": 16, "TIPO ANÚNCIO": 16, "CATÁLOGO": 12, "QTD VENDIDA": 14}
    for col_idx, header in enumerate(rows[0], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(header, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return redirect(PUBLIC_EXPORT_URL, code=302)


@app.route("/export")
@app.route("/export/")
def export_index():
    return render_template("index.html")


@app.route("/api/validate_token", methods=["POST"])
@app.route("/export/api/validate_token", methods=["POST"])
def api_validate_token():
    data = request.json or {}
    token = (data.get("token") or "").strip()
    if not token:
        return jsonify({"valid": False, "error": "Token vazio."}), 400

    user = validate_token(token)
    if user and user.get("id"):
        return jsonify({"valid": True, "nickname": user.get("nickname"), "id": user.get("id")})
    return jsonify({"valid": False, "error": "Token inválido ou expirado."}), 401


@app.route("/api/export", methods=["POST"])
@app.route("/export/api/export", methods=["POST"])
def api_export():
    data     = request.json or {}
    token    = (data.get("token") or "").strip()
    mlb_list = data.get("mlbs", [])

    if not token:
        return jsonify({"success": False, "error": "Token é obrigatório."}), 400
    if not mlb_list:
        return jsonify({"success": False, "error": "Nenhum MLB fornecido."}), 400

    print(f"[EXPORT] {len(mlb_list)} MLBs recebidos")

    try:
        rows, errors = process_mlbs(mlb_list, token)
    except Exception as exc:
        return jsonify({"success": False, "error": f"Erro interno: {str(exc)}"}), 500

    product_count = len(rows) - 1

    if product_count == 0:
        return jsonify({
            "success": False,
            "error": "Nenhum produto válido encontrado.",
            "warnings": errors,
        }), 422

    buf = _write_excel(rows)
    filename = f"Relatorio_MLB_{uuid.uuid4().hex[:8]}.xlsx"

    response = send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["X-Export-Count"]    = str(product_count)
    response.headers["X-Export-Warnings"] = json.dumps(errors, ensure_ascii=False)
    return response


if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 3002))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    print(f"Servidor em http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", debug=debug, port=port)
